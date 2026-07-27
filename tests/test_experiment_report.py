import json

from openpyxl import load_workbook

from coder_review_benchmark.experiment_report import generate_experiment_report


EXPECTED_SHEETS = [
    "Experiment", "Model Summary", "SWE Metrics", "SWE Confusion", "SWE By Generator",
    "SWE By Difficulty", "SWE Samples", "Martian Metrics", "Martian By Repository",
    "Martian By Language", "Martian Judge Errors", "Martian Samples", "Token Usage",
    "Latency", "Run Failures", "Manifests",
]


def test_report_has_classified_excel_sheets_status_and_percentages(tmp_path):
    experiment = tmp_path / "exp"
    run = experiment / "runs" / "swe"
    run.mkdir(parents=True)
    (run / "run_manifest.json").write_text(json.dumps({"experiment_id": "report-exp"}), encoding="utf-8")
    (run / "results.jsonl").write_text(json.dumps({"suite": "swe_review", "status": "completed", "review": {"schema_valid": True}, "elapsed": 0.25}) + "\n", encoding="utf-8")
    metrics = {"suite": "swe_review", "sample_count": 1, "format_completion_rate": 2 / 3, "schema_completion_rate": 1.0, "decision_accuracy_all": 2 / 3, "confusion_matrix": {"approve_resolved": 1}, "by_generator_model": {"g": {"sample_count": 1}}, "by_difficulty": {"easy": {"sample_count": 1}}}
    (run / "metrics.json").write_text(json.dumps(metrics), encoding="utf-8")
    state = {"experiment_id": "report-exp", "status": "completed", "tasks": {"swe": {"suite": "swe_review", "model_profile": "m", "dataset_profile": "d", "status": "completed", "judge_status": "not_applicable", "run_dir": str(run)}}}
    (experiment / "state.json").write_text(json.dumps(state), encoding="utf-8")
    paths = generate_experiment_report(experiment)
    assert load_workbook(paths["excel"], read_only=True).sheetnames == EXPECTED_SHEETS
    markdown = paths["markdown"].read_text(encoding="utf-8")
    assert "Status: **COMPLETED**" in markdown
    suite_markdown = (paths["directory"] / "swe_review" / "report.md").read_text(encoding="utf-8")
    assert "66.67%" in suite_markdown
