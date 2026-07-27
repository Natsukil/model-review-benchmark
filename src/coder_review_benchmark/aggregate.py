from __future__ import annotations

import csv
import json
import re
from datetime import datetime
from pathlib import Path
from typing import Any


SUITES = ("swe_review", "martian", "agentic", "codereviewqa")


def _json(path: Path) -> dict[str, Any]:
    try:
        value = json.loads(path.read_text(encoding="utf-8"))
        return value if isinstance(value, dict) else {}
    except (OSError, json.JSONDecodeError):
        return {}


def _identity(run_dir: Path, manifest: dict[str, Any], metrics: dict[str, Any]) -> tuple[str, str]:
    suite = str(manifest.get("suite") or metrics.get("suite") or "unknown")
    name = run_dir.name
    match = re.match(r"^(\d{8}-\d{6})_(.+)$", name)
    remainder = match.group(2) if match else name
    suffix = f"_{suite}"
    fallback_model = remainder[:-len(suffix)] if remainder.endswith(suffix) else remainder
    return str(manifest.get("model_profile") or fallback_model), suite


def _result_stats(path: Path) -> dict[str, int | float]:
    stats: dict[str, int | float] = {
        "sample_count": 0,
        "completed": 0,
        "malformed_calls": 0,
        "model_elapsed_seconds": 0.0,
        "judge_elapsed_seconds": 0.0,
    }
    if not path.exists():
        return stats
    with path.open(encoding="utf-8") as source:
        for line in source:
            if not line.strip():
                continue
            try:
                row = json.loads(line)
            except json.JSONDecodeError:
                continue
            stats["sample_count"] += 1
            stats["completed"] += row.get("status") == "completed"
            stats["malformed_calls"] += int(row.get("malformed_calls") or 0)
            stats["model_elapsed_seconds"] += float(row.get("elapsed") or 0.0)
            stats["judge_elapsed_seconds"] += float(row.get("judge", {}).get("judge_elapsed") or 0.0)
    return stats


def collect_runs(outputs_dir: Path) -> list[dict[str, Any]]:
    runs: list[dict[str, Any]] = []
    if not outputs_dir.exists():
        return runs
    for metrics_path in sorted(outputs_dir.glob("*/metrics.json")):
        run_dir = metrics_path.parent
        metrics = _json(metrics_path)
        manifest = _json(run_dir / "run_manifest.json")
        model_profile, suite = _identity(run_dir, manifest, metrics)
        stats = _result_stats(run_dir / "results.jsonl")
        timestamp = run_dir.name[:15] if re.match(r"^\d{8}-\d{6}", run_dir.name) else ""
        confusion = metrics.get("confusion_matrix") if isinstance(metrics.get("confusion_matrix"), dict) else {}
        row: dict[str, Any] = {
            "run_id": run_dir.name,
            "timestamp": timestamp,
            "model_profile": model_profile,
            "model_name": manifest.get("model_name") or model_profile,
            "suite": suite,
            "profile": manifest.get("profile") or "legacy",
            "evaluation_version": manifest.get("evaluation_version") or (
                "batch-one-to-one-v1" if "judge_calls" in metrics else "legacy"
            ),
            "sample_count": stats["sample_count"],
            "completed": stats["completed"],
            "completion_rate": metrics.get("completion_rate"),
            "malformed_calls": stats["malformed_calls"],
            "decision_accuracy": metrics.get("decision_accuracy"),
            "average_findings": metrics.get("average_findings"),
            "approved_resolved": confusion.get("approved_resolved"),
            "approved_unresolved": confusion.get("approved_unresolved"),
            "rejected_resolved": confusion.get("rejected_resolved"),
            "rejected_unresolved": confusion.get("rejected_unresolved"),
            "precision": metrics.get("precision"),
            "recall": metrics.get("recall"),
            "f1": metrics.get("f1"),
            "tp": metrics.get("tp"),
            "fp": metrics.get("fp"),
            "fn": metrics.get("fn"),
            "judge_calls": metrics.get("judge_calls"),
            "judge_errors": metrics.get("judge_errors"),
            "resolve_rate": metrics.get("resolve_rate"),
            "evaluated": metrics.get("evaluated"),
            "resolved": metrics.get("resolved"),
            "model_elapsed_seconds": metrics.get("model_elapsed_seconds", stats["model_elapsed_seconds"]),
            "judge_elapsed_seconds": metrics.get("judge_elapsed_seconds", stats["judge_elapsed_seconds"]),
            "wall_elapsed_seconds": metrics.get("wall_elapsed_seconds"),
            "run_dir": str(run_dir.resolve()),
            "by_generator_model": metrics.get("by_generator_model") or {},
            "by_difficulty": metrics.get("by_difficulty") or {},
        }
        if row["completion_rate"] is None:
            row["completion_rate"] = stats["completed"] / stats["sample_count"] if stats["sample_count"] else 0.0
        runs.append(row)
    return runs


def select_comparison_runs(runs: list[dict[str, Any]], profile: str | None = None) -> list[dict[str, Any]]:
    selected: dict[tuple[str, str], dict[str, Any]] = {}
    for row in runs:
        if profile is not None and row.get("profile") != profile:
            continue
        key = (str(row["model_profile"]), str(row["suite"]))
        rank = (int(row["sample_count"]), float(row["completion_rate"] or 0.0), str(row["timestamp"]))
        current = selected.get(key)
        current_rank = (
            int(current["sample_count"]),
            float(current["completion_rate"] or 0.0),
            str(current["timestamp"]),
        ) if current else None
        if current_rank is None or rank > current_rank:
            selected[key] = row
    return sorted(selected.values(), key=lambda row: (str(row["model_profile"]), str(row["suite"])))


def _percent(value: Any) -> str:
    return "—" if value is None else f"{float(value):.2%}"


def _minutes(value: Any) -> str:
    return "—" if value is None else f"{float(value):.1f} 分钟"


def _comparison_rows(selected: list[dict[str, Any]]) -> list[dict[str, Any]]:
    indexed = {(row["model_profile"], row["suite"]): row for row in selected}
    result = []
    for model in sorted({str(row["model_profile"]) for row in selected}):
        swe = indexed.get((model, "swe_review"), {})
        martian = indexed.get((model, "martian"), {})
        agentic = indexed.get((model, "agentic"), {})
        swe_seconds = float(swe.get("model_elapsed_seconds") or 0.0)
        martian_seconds = float(martian.get("model_elapsed_seconds") or 0.0)
        result.append({
            "model": model,
            "swe_samples": swe.get("sample_count"),
            "swe_completion": swe.get("completion_rate"),
            "swe_accuracy": swe.get("decision_accuracy"),
            "swe_defect_recall": (
                swe.get("rejected_unresolved") / (swe.get("rejected_unresolved") + swe.get("approved_unresolved"))
                if swe.get("rejected_unresolved") is not None and swe.get("approved_unresolved") is not None
                and swe.get("rejected_unresolved") + swe.get("approved_unresolved") else None
            ),
            "martian_samples": martian.get("sample_count"),
            "martian_completion": martian.get("completion_rate"),
            "martian_precision": martian.get("precision"),
            "martian_recall": martian.get("recall"),
            "martian_f1": martian.get("f1"),
            "review_model_minutes": (swe_seconds + martian_seconds) / 60 if swe or martian else None,
            "agentic_samples": agentic.get("sample_count"),
            "agentic_resolve_rate": agentic.get("resolve_rate"),
        })
    return result


def write_markdown(path: Path, selected: list[dict[str, Any]], profile: str | None) -> None:
    comparison = _comparison_rows(selected)
    lines = [
        "# 模型评测汇总",
        "",
        f"生成时间：{datetime.now().astimezone().isoformat(timespec='seconds')}",
        "",
        f"对比 Profile：`{profile or '全部'}`",
        "",
        "选取规则：同一模型和套件优先选择样本数最多的运行，其次选择完成率最高、时间最新的运行。",
        "",
        "| 模型 | SWE 样本 | SWE 准确率 | 缺陷拒绝率 | Martian 样本 | Martian Precision | Martian Recall | Martian F1 | Review 模型耗时 | Agentic Resolve |",
        "|---|---:|---:|---:|---:|---:|---:|---:|---:|---:|",
    ]
    for row in comparison:
        lines.append(
            f"| {row['model']} | {row['swe_samples'] or '—'} | {_percent(row['swe_accuracy'])} | "
            f"{_percent(row['swe_defect_recall'])} | {row['martian_samples'] or '—'} | "
            f"{_percent(row['martian_precision'])} | {_percent(row['martian_recall'])} | "
            f"{_percent(row['martian_f1'])} | {_minutes(row['review_model_minutes'])} | "
            f"{_percent(row['agentic_resolve_rate'])} |"
        )
    lines.extend(["", "## 被选中的运行", "", "| 模型 | 套件 | Profile | 样本 | 完成率 | 运行目录 |", "|---|---|---|---:|---:|---|"])
    for row in selected:
        lines.append(
            f"| {row['model_profile']} | {row['suite']} | {row['profile']} | {row['sample_count']} | "
            f"{_percent(row['completion_rate'])} | `{row['run_id']}` |"
        )
    path.write_text("\n".join(lines) + "\n", encoding="utf-8")


def write_csv(path: Path, runs: list[dict[str, Any]]) -> None:
    fields = [key for key in runs[0] if key not in {"by_generator_model", "by_difficulty"}] if runs else []
    with path.open("w", newline="", encoding="utf-8-sig") as output:
        writer = csv.DictWriter(output, fieldnames=fields)
        if fields:
            writer.writeheader()
            writer.writerows({key: row.get(key) for key in fields} for row in runs)


def write_excel(path: Path, runs: list[dict[str, Any]], selected: list[dict[str, Any]]) -> None:
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill
    except ImportError as exc:
        raise RuntimeError("Excel 汇总需要安装 openpyxl：pip install openpyxl") from exc

    workbook = Workbook()
    workbook.remove(workbook.active)
    sheets = [
        ("模型对比", _comparison_rows(selected)),
        ("选中运行", [{k: v for k, v in row.items() if k not in {"by_generator_model", "by_difficulty"}} for row in selected]),
        ("全部运行", [{k: v for k, v in row.items() if k not in {"by_generator_model", "by_difficulty"}} for row in runs]),
    ]
    breakdown = []
    for row in selected:
        for dimension in ("by_generator_model", "by_difficulty"):
            for value, metrics in row.get(dimension, {}).items():
                breakdown.append({"model": row["model_profile"], "suite": row["suite"], "dimension": dimension, "value": value, **metrics})
    sheets.append(("SWE拆分", breakdown))

    for title, rows in sheets:
        sheet = workbook.create_sheet(title)
        if not rows:
            sheet.append(["暂无数据"])
            continue
        headers = list(rows[0])
        sheet.append(headers)
        for row in rows:
            sheet.append([row.get(header) for header in headers])
        sheet.freeze_panes = "A2"
        sheet.auto_filter.ref = sheet.dimensions
        for cell in sheet[1]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill("solid", fgColor="1F4E78")
        for column in sheet.columns:
            letter = column[0].column_letter
            sheet.column_dimensions[letter].width = min(60, max(12, max(len(str(cell.value or "")) for cell in column) + 2))
        for row_cells in sheet.iter_rows(min_row=2):
            for cell in row_cells:
                if isinstance(cell.value, float) and any(word in str(sheet.cell(1, cell.column).value) for word in ("rate", "accuracy", "precision", "recall", "f1", "completion")):
                    cell.number_format = "0.00%"
    workbook.save(path)


def summarize_runs(outputs_dir: Path, output_dir: Path, profile: str | None = "review-hour") -> dict[str, Path]:
    runs = collect_runs(outputs_dir)
    if not runs:
        raise RuntimeError(f"没有在 {outputs_dir} 中找到可汇总的 metrics.json")
    selected = select_comparison_runs(runs, profile)
    if not selected:
        raise RuntimeError(f"没有找到 profile={profile!r} 的可比较运行")
    output_dir.mkdir(parents=True, exist_ok=True)
    paths = {
        "markdown": output_dir / "benchmark_summary.md",
        "csv": output_dir / "benchmark_all_runs.csv",
        "excel": output_dir / "benchmark_summary.xlsx",
    }
    write_markdown(paths["markdown"], selected, profile)
    write_csv(paths["csv"], runs)
    write_excel(paths["excel"], runs, selected)
    return paths
