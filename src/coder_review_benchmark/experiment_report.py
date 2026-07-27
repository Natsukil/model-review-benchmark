from __future__ import annotations

import csv
import json
from collections import Counter
from pathlib import Path
from typing import Any


def _read_json(path: Path) -> dict[str, Any]:
    value = json.loads(path.read_text(encoding="utf-8"))
    return value if isinstance(value, dict) else {}


def _read_jsonl(path: Path) -> list[dict[str, Any]]:
    if not path.exists():
        return []
    return [json.loads(line) for line in path.read_text(encoding="utf-8").splitlines() if line.strip()]


def _common(rows: list[dict[str, Any]]) -> dict[str, Any]:
    return {
        "sample_count": len(rows),
        "prompt_tokens": sum(int(r.get("prompt_tokens") or 0) for r in rows),
        "completion_tokens": sum(int(r.get("completion_tokens") or 0) for r in rows),
        "total_tokens": sum(int(r.get("total_tokens") or 0) for r in rows),
        "finish_reason_counts": dict(Counter(str(r.get("finish_reason") or "unknown") for r in rows)),
        "truncated_samples": sum(bool(r.get("truncated")) for r in rows),
        "latency_seconds": sum(float(r.get("elapsed") or 0.0) for r in rows),
        "run_errors": sum(r.get("status") == "error" for r in rows),
    }


def _flatten(value: Any) -> str | int | float | bool | None:
    return json.dumps(value, ensure_ascii=False, sort_keys=True) if isinstance(value, (dict, list)) else value


def _write_csv(path: Path, rows: list[dict[str, Any]]) -> None:
    fields = sorted({key for row in rows for key in row}) if rows else ["status"]
    with path.open("w", newline="", encoding="utf-8-sig") as output:
        writer = csv.DictWriter(output, fieldnames=fields, extrasaction="ignore")
        writer.writeheader()
        writer.writerows({key: _flatten(row.get(key)) for key in fields} for row in rows)


def _write_excel(path: Path, summary_rows: list[dict[str, Any]], sample_rows: list[dict[str, Any]]) -> None:
    from openpyxl import Workbook

    workbook = Workbook()
    workbook.remove(workbook.active)
    for title, rows in (("metrics", summary_rows), ("per_sample", sample_rows)):
        sheet = workbook.create_sheet(title[:31])
        fields = sorted({key for row in rows for key in row}) if rows else ["status"]
        sheet.append(fields)
        for row in rows:
            sheet.append([_flatten(row.get(field)) for field in fields])
        sheet.freeze_panes = "A2"
    workbook.save(path)


def _row_status(row: dict[str, Any]) -> str:
    if row.get("status") != "completed":
        return "failed"
    if row.get("suite") == "martian":
        return str(row.get("judge_status") or "failed")
    return "completed"


def _report_status(rows: list[dict[str, Any]]) -> str:
    statuses = [_row_status(row) for row in rows]
    if "partial" in statuses:
        return "partial"
    if "failed" in statuses and "completed" in statuses:
        return "partial"
    if statuses and all(status == "failed" for status in statuses):
        return "failed"
    return "completed"


def _breakdowns(rows: list[dict[str, Any]], field: str, label: str) -> list[dict[str, Any]]:
    output: list[dict[str, Any]] = []
    for row in rows:
        groups = row.get(field) if isinstance(row.get(field), dict) else {}
        for name, metrics in groups.items():
            output.append({"model_profile": row.get("model_profile"), "dataset_profile": row.get("dataset_profile"), label: name, **(metrics if isinstance(metrics, dict) else {})})
    return output


def _judge_errors(samples: list[dict[str, Any]]) -> list[dict[str, Any]]:
    output: list[dict[str, Any]] = []
    for sample in samples:
        judge = sample.get("judge") if isinstance(sample.get("judge"), dict) else {}
        errors = judge.get("errors", [])
        for error in errors:
            output.append({"task_id": sample.get("task_id"), "model_profile": sample.get("model_profile"), "sample_id": sample.get("sample_id"), "pr_url": sample.get("pr_url"), "raw_response": judge.get("raw_response"), "request_attempts": judge.get("request_attempts"), "finish_reason": judge.get("finish_reason"), "elapsed": judge.get("elapsed", judge.get("judge_elapsed")), "schema_error": judge.get("schema_error"), **(error if isinstance(error, dict) else {"error": str(error)})})
    return output


def _write_complete_excel(path: Path, experiment_id: str, report_status: str, state: dict[str, Any], summaries: list[dict[str, Any]], samples: list[dict[str, Any]], manifests: list[dict[str, Any]], judge_manifest: dict[str, Any] | None) -> None:
    from openpyxl import Workbook

    swe = [row for row in summaries if row.get("suite") == "swe_review"]
    martian = [row for row in summaries if row.get("suite") == "martian"]
    swe_samples = [row for row in samples if row.get("suite") == "swe_review"]
    martian_samples = [row for row in samples if row.get("suite") == "martian"]
    confusion = [{"model_profile": row.get("model_profile"), "dataset_profile": row.get("dataset_profile"), **(row.get("confusion_matrix") or {})} for row in swe]
    token_usage = [{key: row.get(key) for key in ("task_id", "model_profile", "suite", "index", "prompt_tokens", "completion_tokens", "total_tokens", "request_attempts", "finish_reason", "truncated")} for row in samples]
    latency = [{key: row.get(key) for key in ("task_id", "model_profile", "suite", "index", "elapsed")} | {"judge_elapsed": (row.get("judge") or {}).get("judge_elapsed") if isinstance(row.get("judge"), dict) else None} for row in samples]
    failures = [{"task_id": task_id, "suite": task.get("suite"), "model_profile": task.get("model_profile"), "status": task.get("status"), "error": task.get("error"), "judge_status": task.get("judge_status"), "judge_error": task.get("judge_error")} for task_id, task in state.get("tasks", {}).items() if task.get("status") != "completed" or (task.get("suite") == "martian" and task.get("judge_status") != "completed")]
    manifest_rows = [{"kind": "run", **manifest} for manifest in manifests]
    if judge_manifest:
        manifest_rows.append({"kind": "judge", **judge_manifest})
    sheets = [
        ("Experiment", [{"experiment_id": experiment_id, "status": report_status, "created_at": state.get("created_at"), "finished_at": state.get("finished_at"), "config_sha256": state.get("config_sha256")}]),
        ("Model Summary", summaries),
        ("SWE Metrics", swe),
        ("SWE Confusion", confusion),
        ("SWE By Generator", _breakdowns(swe, "by_generator_model", "generator")),
        ("SWE By Difficulty", _breakdowns(swe, "by_difficulty", "difficulty")),
        ("SWE Samples", swe_samples),
        ("Martian Metrics", martian),
        ("Martian By Repository", _breakdowns(martian, "per_repository_metrics", "repository")),
        ("Martian By Language", _breakdowns(martian, "by_language", "language")),
        ("Martian Judge Errors", _judge_errors(martian_samples)),
        ("Martian Samples", martian_samples),
        ("Token Usage", token_usage),
        ("Latency", latency),
        ("Run Failures", failures),
        ("Manifests", manifest_rows),
    ]
    workbook = Workbook()
    workbook.remove(workbook.active)
    for title, rows in sheets:
        sheet = workbook.create_sheet(title)
        fields = sorted({key for row in rows for key in row}) if rows else ["status"]
        sheet.append(fields)
        for row in rows:
            sheet.append([_flatten(row.get(field)) for field in fields])
        sheet.freeze_panes = "A2"
    workbook.save(path)


def _format_metric(key: str, value: Any) -> str:
    if value is None:
        return ""
    percentage = any(part in key.lower() for part in ("rate", "accuracy", "precision", "recall", "_f1", "completion"))
    if percentage and isinstance(value, (int, float)):
        return f"{float(value):.2%}"
    if isinstance(value, float):
        return f"{value:.3f}"
    return str(value)


def generate_experiment_report(experiment_dir: Path, output_dir: Path | None = None) -> dict[str, Path]:
    state = _read_json(experiment_dir / "state.json")
    experiment_id = str(state.get("experiment_id") or experiment_dir.name)
    output_dir = output_dir or experiment_dir / "reports"
    output_dir.mkdir(parents=True, exist_ok=True)
    suites: dict[str, dict[str, Any]] = {}
    all_summary: list[dict[str, Any]] = []
    all_samples: list[dict[str, Any]] = []
    manifests: list[dict[str, Any]] = []
    for task_id, task in state.get("tasks", {}).items():
        suite = str(task.get("suite", "unknown"))
        bucket = suites.setdefault(suite, {"runs": [], "samples": []})
        run_dir_value = task.get("run_dir")
        entry = {"task_id": task_id, "status": task.get("status"), "error": task.get("error"), "judge_status": task.get("judge_status"), "judge_error": task.get("judge_error"), "model_profile": task.get("model_profile"), "dataset_profile": task.get("dataset_profile")}
        if run_dir_value:
            run_dir = Path(run_dir_value)
            manifest_path = run_dir / "run_manifest.json"
            metrics_path = run_dir / "metrics.json"
            if manifest_path.exists():
                manifest = _read_json(manifest_path)
                if manifest.get("experiment_id") != experiment_id:
                    entry["status"] = "manifest_mismatch"
                    entry["error"] = "run belongs to a different experiment_id"
                else:
                    manifests.append(manifest)
            metrics = _read_json(metrics_path) if metrics_path.exists() else {}
            rows = _read_jsonl(run_dir / "results.jsonl")
            entry.update(metrics)
            entry.update(_common(rows))
            for row in rows:
                sample = {"task_id": task_id, "model_profile": task.get("model_profile"), "dataset_profile": task.get("dataset_profile"), **row}
                bucket["samples"].append(sample)
                all_samples.append(sample)
        entry["effective_status"] = _row_status({"suite": suite, **entry})
        bucket["runs"].append(entry)
        all_summary.append({"suite": suite, **entry})

    for suite, data in suites.items():
        suite_dir = output_dir / suite
        suite_dir.mkdir(parents=True, exist_ok=True)
        suite_status = _report_status([{"suite": suite, **row} for row in data["runs"]])
        payload = {"experiment_id": experiment_id, "status": suite_status, "suite": suite, "runs": data["runs"], "per_sample": data["samples"]}
        (suite_dir / "report.json").write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")
        _write_csv(suite_dir / "metrics.csv", data["runs"])
        _write_csv(suite_dir / "per_sample.csv", data["samples"])
        _write_excel(suite_dir / "report.xlsx", data["runs"], data["samples"])
        if suite == "swe_review":
            fields = [("format_completion_rate", "Format"), ("schema_completion_rate", "Schema"), ("decision_accuracy_all", "Acc all"), ("decision_accuracy_valid", "Acc valid"), ("balanced_accuracy", "Balanced"), ("defect_recall", "Recall"), ("false_acceptance_rate", "FAR"), ("false_rejection_rate", "FRR"), ("MCC", "MCC")]
        else:
            fields = [("tp", "TP"), ("fp", "FP"), ("fn", "FN"), ("micro_precision", "Micro P"), ("micro_recall", "Micro R"), ("micro_f1", "Micro F1"), ("macro_precision", "Macro P"), ("macro_recall", "Macro R"), ("macro_f1", "Macro F1"), ("average_findings", "Avg findings"), ("zero_finding_prs", "Zero findings"), ("judge_errors", "Judge errors")]
        headers = [label for _, label in fields]
        lines = [f"# {suite} report", "", f"Experiment: `{experiment_id}`", "", f"Status: **{suite_status.upper()}**", "", "| Model | Dataset | Status | Samples | " + " | ".join(headers) + " |", "|---|---|---|---:|" + "---:|" * len(fields)]
        for row in data["runs"]:
            values = [_format_metric(key, row.get(key)) for key, _ in fields]
            lines.append(f"| {row.get('model_profile')} | {row.get('dataset_profile')} | {_row_status({'suite': suite, **row})} | {row.get('sample_count', 0)} | " + " | ".join(values) + " |")
        lines.extend(["", "Breakdowns and complete per-sample records are included in `report.json`, `metrics.csv`, `per_sample.csv`, and `report.xlsx`."])
        (suite_dir / "report.md").write_text("\n".join(lines) + "\n", encoding="utf-8")

    judge_manifest_path = experiment_dir / "judge_manifest.json"
    judge_manifest = _read_json(judge_manifest_path) if judge_manifest_path.exists() else None
    report_status = _report_status(all_summary)
    complete = {"experiment_id": experiment_id, "status": report_status, "state": state, "manifests": manifests, "judge_manifest": judge_manifest, "runs": all_summary, "suites": {suite: {"run_count": len(data["runs"]), "sample_count": len(data["samples"])} for suite, data in suites.items()}}
    json_path = output_dir / "report.json"
    json_path.write_text(json.dumps(complete, ensure_ascii=False, indent=2), encoding="utf-8")
    _write_csv(output_dir / "report.csv", all_summary)
    _write_complete_excel(output_dir / "report.xlsx", experiment_id, report_status, state, all_summary, all_samples, manifests, judge_manifest)
    md_lines = ["# Model Review Benchmark experiment", "", f"Experiment: `{experiment_id}`", "", f"Status: **{report_status.upper()}**", "", "| Suite | Model | Dataset | Status | Samples | Errors |", "|---|---|---|---|---:|---:|"]
    for row in all_summary:
        md_lines.append(f"| {row.get('suite')} | {row.get('model_profile')} | {row.get('dataset_profile')} | {_row_status(row)} | {row.get('sample_count', 0)} | {row.get('run_errors', 0)} |")
    markdown_path = output_dir / "report.md"
    markdown_path.write_text("\n".join(md_lines) + "\n", encoding="utf-8")
    return {"markdown": markdown_path, "json": json_path, "csv": output_dir / "report.csv", "excel": output_dir / "report.xlsx", "directory": output_dir}
